import io
import os
import json
import tempfile
import msoffcrypto
import pandas as pd
import uvicorn
from difflib import SequenceMatcher
from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI(title="Payroll Hours Validator API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

ACCOUNTING_FORMAT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'

def clean_id(val):
    if pd.isna(val): return ""
    val_str = str(val).strip()
    if val_str.endswith(".0"): val_str = val_str[:-2]
    return val_str.replace(" ", "")

def normalize_text(text):
    return str(text).lower().replace("sum of ", "").replace("  ", " ").strip()

def get_similarity(a, b):
    return SequenceMatcher(None, normalize_text(a), normalize_text(b)).ratio()

def smart_match_columns(main_cols, dr2_cols, threshold=0.55):
    matched_map = {}
    ambiguous_cols = []
    for m_col in main_cols:
        best_match, highest_score = None, 0.0
        for d_col in dr2_cols:
            score = get_similarity(m_col, d_col)
            norm_m, norm_d = normalize_text(m_col), normalize_text(d_col)
            if norm_m in norm_d or norm_d in norm_m:
                score = max(score, 0.85)
            if score > highest_score:
                highest_score, best_match = score, d_col
        if highest_score >= threshold:
            matched_map[m_col] = best_match
        else:
            ambiguous_cols.append(m_col)
            matched_map[m_col] = "-- Skip / Do Not Compare --"
    return matched_map, ambiguous_cols

def get_target_sheet(excel_file, sheet_identifier):
    if not sheet_identifier:
        return excel_file.sheet_names[0]
    if isinstance(sheet_identifier, int) or str(sheet_identifier).isdigit():
        idx = int(sheet_identifier)
        return excel_file.sheet_names[idx] if idx < len(excel_file.sheet_names) else excel_file.sheet_names[0]
    target_clean = str(sheet_identifier).strip().lower()
    matched = [s for s in excel_file.sheet_names if s.strip().lower() == target_clean]
    return matched[0] if matched else excel_file.sheet_names[0]

def load_file_robust(file_bytes, file_name, password, sheet_identifier):
    if file_name.lower().endswith('.csv'):
        try: return pd.read_csv(io.BytesIO(file_bytes))
        except Exception:
            return pd.read_csv(io.BytesIO(file_bytes), encoding='latin1')

    engines = ["openpyxl", "pyxlsb", "xlrd"]
    for engine in engines:
        try:
            excel_file = pd.ExcelFile(io.BytesIO(file_bytes), engine=engine)
            sheet = get_target_sheet(excel_file, sheet_identifier)
            df = pd.read_excel(excel_file, sheet_name=sheet)
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except Exception: continue

    if password:
        decrypted_stream = io.BytesIO()
        try:
            office_file = msoffcrypto.OfficeFile(io.BytesIO(file_bytes))
            office_file.load_key(password=password)
            office_file.decrypt(decrypted_stream)
        except Exception:
            raise ValueError("Incorrect password or failed to decrypt the file.")

        decrypted_stream.seek(0)
        for engine in engines:
            try:
                excel_file = pd.ExcelFile(decrypted_stream, engine=engine)
                sheet = get_target_sheet(excel_file, sheet_identifier)
                df = pd.read_excel(excel_file, sheet_name=sheet)
                df.columns = [str(c).strip() for c in df.columns]
                return df
            except Exception: continue

    raise ValueError("Unable to read the file. Please check the file format or password.")

def get_best_dr2_key(df_dr2):
    for c in df_dr2.columns:
        c_low = str(c).lower()
        if any(k in c_low for k in ["emp id", "employee id", "empid", "emp_id", "row label"]):
            return c
    return next((c for c in df_dr2.columns if any(k in str(c).lower() for k in ["id", "emp"])), df_dr2.columns[0])

def parse_numeric_val(val):
    if pd.isna(val):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    val_str = str(val).replace(',', '').strip()
    if not val_str or val_str == '-':
        return 0.0
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def process_single_sheet_dataframe(df_main, df_dr2, key_main, key_dr2, final_column_map):
    df_main[key_main] = df_main[key_main].apply(clean_id)
    df_dr2[key_dr2] = df_dr2[key_dr2].apply(clean_id)

    final_cols, display_header_map = [], {}

    for col in df_main.columns:
        if col == key_main:
            final_cols.append(col)
            display_header_map[col] = col
            continue

        final_cols.append(col)
        display_header_map[col] = col
        dr2_target_col = final_column_map.get(col)

        if dr2_target_col and dr2_target_col != "-- Skip / Do Not Compare --":
            dr2_val_col, chk_col, rem_col = f"__DR2_{col}", f"__CHK_{col}", f"__REM_{col}"
            final_cols.extend([dr2_val_col, chk_col, rem_col])
            display_header_map[dr2_val_col] = "DR2"
            display_header_map[chk_col] = "CHECKER"
            display_header_map[rem_col] = "REMARKS"

            df_dr2_clean = df_dr2.copy()
            
            # Determine if this column is numeric or string
            is_numeric_col = True
            if col in df_main.columns:
                sample_vals = df_main[col].dropna().head(10)
                numeric_count = 0
                for v in sample_vals:
                    try:
                        float(str(v).replace(',', '').strip())
                        numeric_count += 1
                    except ValueError:
                        pass
                if len(sample_vals) > 0 and (numeric_count / len(sample_vals)) < 0.5:
                    is_numeric_col = False

            if dr2_target_col in df_dr2_clean.columns:
                if is_numeric_col:
                    df_dr2_clean[dr2_target_col] = df_dr2_clean[dr2_target_col].apply(parse_numeric_val)
                    dr2_grouped = df_dr2_clean.groupby(key_dr2)[dr2_target_col].sum().to_dict()
                else:
                    df_dr2_clean[dr2_target_col] = df_dr2_clean[dr2_target_col].apply(lambda x: "" if pd.isna(x) else str(x).strip())
                    dr2_grouped = df_dr2_clean.groupby(key_dr2)[dr2_target_col].first().to_dict()
            else:
                dr2_grouped = {}

            dr2_vals, chk_vals, rem_vals = [], [], []
            for _, row in df_main.iterrows():
                emp_id = row[key_main]
                
                if is_numeric_col:
                    main_val = parse_numeric_val(row[col])
                    if emp_id in dr2_grouped:
                        dr2_val = dr2_grouped[emp_id]
                        diff = round(main_val - dr2_val, 4)
                        dr2_vals.append(dr2_val)
                        chk_vals.append(diff)
                        rem_vals.append("OK; NO VARIANCE" if abs(diff) < 0.0001 else "VARIANCE")
                    else:
                        dr2_vals.append(0.0)
                        chk_vals.append(main_val)
                        rem_vals.append("NOT IN DR")
                else:
                    main_val = "" if pd.isna(row[col]) else str(row[col]).strip()
                    if emp_id in dr2_grouped:
                        dr2_val = dr2_grouped[emp_id]
                        dr2_vals.append(dr2_val)
                        
                        is_equal = (main_val.lower() == dr2_val.lower())
                        chk_vals.append(True if is_equal else False)
                        rem_vals.append("OK; MATCH" if is_equal else "MISMATCH")
                    else:
                        dr2_vals.append("")
                        chk_vals.append(False)
                        rem_vals.append("NOT IN DR")

            df_main[dr2_val_col] = dr2_vals
            df_main[chk_col] = chk_vals
            df_main[rem_col] = rem_vals

    df_final = df_main[final_cols]
    return df_final, final_cols, display_header_map

@app.get("/", response_class=HTMLResponse)
async def serve_ui():
    if not os.path.exists("index.html"):
        raise HTTPException(status_code=404, detail="index.html not found.")
    with open("index.html", "r", encoding="utf-8") as f:
        return f.read()

@app.post("/get-sheets")
async def get_sheets(
    file: UploadFile = File(...),
    password: str = Form("")
):
    try:
        file_bytes = await file.read()
        
        if file.filename.lower().endswith('.csv'):
            return JSONResponse({"sheet_names": ["Sheet1"]})

        engines = ["openpyxl", "pyxlsb", "xlrd"]
        sheet_names = []
        for engine in engines:
            try:
                excel_file = pd.ExcelFile(io.BytesIO(file_bytes), engine=engine)
                sheet_names = excel_file.sheet_names
                break
            except Exception:
                continue
        
        if not sheet_names and password:
            decrypted_stream = io.BytesIO()
            office_file = msoffcrypto.OfficeFile(io.BytesIO(file_bytes))
            office_file.load_key(password=password)
            office_file.decrypt(decrypted_stream)
            decrypted_stream.seek(0)
            for engine in engines:
                try:
                    excel_file = pd.ExcelFile(decrypted_stream, engine=engine)
                    sheet_names = excel_file.sheet_names
                    break
                except Exception:
                    continue

        if not sheet_names:
            raise ValueError("No readable sheets found. Please check the file format or password.")

        return JSONResponse({"sheet_names": sheet_names})
    except Exception as e:
        print(f"Error fetching sheets: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Failed to read sheets: {str(e)}")

@app.post("/analyze-mapping")
async def analyze_mapping(
    payroll_inputs: UploadFile = File(...),
    payroll_masterfile: UploadFile = File(...),
    password: str = Form(""),
    sheet_name: str = Form(""),
    employee_id_col: str = Form("")
):
    try:
        main_bytes = await payroll_inputs.read()
        dr2_bytes = await payroll_masterfile.read()

        df_main = load_file_robust(main_bytes, payroll_inputs.filename, password, sheet_name)
        df_dr2 = load_file_robust(dr2_bytes, payroll_masterfile.filename, password, 0)

        if employee_id_col and employee_id_col in df_main.columns:
            key_main = employee_id_col
        else:
            key_main = next((c for c in df_main.columns if any(k in str(c).lower() for k in ["row label", "id", "emp", "code"])), df_main.columns[0])

        key_dr2 = get_best_dr2_key(df_dr2)

        data_cols = [c for c in df_main.columns if c != key_main]
        matched_map, ambiguous_cols = smart_match_columns(data_cols, df_dr2.columns)

        return JSONResponse({
            "main_columns": list(df_main.columns),
            "dr2_columns": list(df_dr2.columns),
            "data_cols": data_cols,
            "detected_key_main": key_main,
            "detected_key_dr2": key_dr2,
            "initial_mapping": matched_map,
            "ambiguous_cols": ambiguous_cols
        })
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/validate-mapped")
async def validate_mapped(
    payroll_inputs: UploadFile = File(...),
    payroll_masterfile: UploadFile = File(...),
    password: str = Form(""),
    sheet_name: str = Form(""),
    employee_id_col: str = Form(""),
    master_id_col: str = Form(""),
    mapping_json: str = Form("{}"),
    all_mappings_json: str = Form("{}"),
    all_id_keys_json: str = Form("{}")
):
    try:
        main_bytes = await payroll_inputs.read()
        dr2_bytes = await payroll_masterfile.read()

        df_dr2 = load_file_robust(dr2_bytes, payroll_masterfile.filename, password, 0)
        key_dr2 = master_id_col if master_id_col in df_dr2.columns else get_best_dr2_key(df_dr2)

        all_mappings = {}
        if all_mappings_json and all_mappings_json != "{}":
            try:
                all_mappings = json.loads(all_mappings_json)
            except Exception:
                all_mappings = {}

        all_id_keys = {}
        if all_id_keys_json and all_id_keys_json != "{}":
            try:
                all_id_keys = json.loads(all_id_keys_json)
            except Exception:
                all_id_keys = {}

        if not all_mappings:
            target_s = sheet_name if sheet_name else ""
            single_map = json.loads(mapping_json) if mapping_json else {}
            all_mappings = {target_s: single_map}

        temp_dir = tempfile.gettempdir()
        output_filepath = os.path.join(temp_dir, "Payroll_Validated_Report.xlsx")

        with pd.ExcelWriter(output_filepath, engine="openpyxl") as writer:
            for s_name, tab_mapping in all_mappings.items():
                if isinstance(tab_mapping, dict) and tab_mapping.get("__SKIPPED__"):
                    continue

                try:
                    df_main = load_file_robust(main_bytes, payroll_inputs.filename, password, s_name if s_name else 0)
                except Exception:
                    continue

                if df_main.empty:
                    continue

                tab_specific_id = all_id_keys.get(s_name)
                if tab_specific_id and tab_specific_id in df_main.columns:
                    key_main = tab_specific_id
                elif employee_id_col and employee_id_col in df_main.columns:
                    key_main = employee_id_col
                else:
                    key_main = next((c for c in df_main.columns if any(k in str(c).lower() for k in ["row label", "id", "emp", "code"])), df_main.columns[0])

                df_final, final_cols, display_header_map = process_single_sheet_dataframe(
                    df_main, df_dr2, key_main, key_dr2, tab_mapping
                )

                sheet_label = str(s_name).strip() if s_name else "Validated"
                if not sheet_label or sheet_label == "None":
                    sheet_label = "Validated"

                df_final.to_excel(writer, sheet_name=sheet_label, index=False)
                ws = writer.sheets[sheet_label]

                thin_border = Border(
                    left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'),
                    top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3')
                )

                header_display_names = [display_header_map[c] for c in final_cols]
                for col_idx, text in enumerate(header_display_names, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value, cell.border = text, thin_border
                    if text in ["DR2", "CHECKER", "REMARKS"]:
                        cell.fill = PatternFill(start_color="D4AC0D", end_color="D4AC0D", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="000000")
                    else:
                        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                key_col_index = final_cols.index(key_main) + 1 if key_main in final_cols else 1
                for row_idx in range(2, ws.max_row + 1):
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font, cell.border = Font(name="Calibri", size=10), thin_border

                        if col_idx == key_col_index:
                            cell.number_format, cell.alignment = '@', Alignment(horizontal="left", vertical="center")
                            continue

                        if isinstance(cell.value, bool):
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                            continue

                        if isinstance(cell.value, (int, float)):
                            cell.number_format, cell.alignment = ACCOUNTING_FORMAT, Alignment(horizontal="right", vertical="center")
                        elif cell.value in ["-", "#N/A"]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                for col in ws.columns:
                    max_len, col_letter = 0, get_column_letter(col[0].column)
                    for cell in col:
                        val = str(cell.value or '')
                        if cell.number_format == ACCOUNTING_FORMAT and isinstance(cell.value, (int, float)):
                            val = f"{cell.value:,.2f}"
                        if len(val) > max_len: max_len = len(val)
                    ws.column_dimensions[col_letter].width = max(max_len + 4, 11)

        return FileResponse(output_filepath, filename="Payroll_Validated_Report.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

if __name__ == "__main__":
    uvicorn.run("app:app", host="0.0.0.0", port=7860, reload=True)